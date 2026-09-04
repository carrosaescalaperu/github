"""
================================================================================
 SCRAPER — Catálogo mayorista modelcarswholesale.com  →  Excel con miniaturas
================================================================================

CONTEXTO Y ADVERTENCIA IMPORTANTE (leer antes de usar):
--------------------------------------------------------------------------------
Al intentar inspeccionar la web para construir este script, la petición fue
bloqueada por un sistema anti-bot (probablemente Cloudflare u otro WAF). Esto
significa dos cosas:

  1. NO pude confirmar los selectores CSS reales (nombres de clases/IDs) de
     las tarjetas de producto, precio, SKU, etc. Los que ves abajo son los
     patrones MÁS COMUNES en catálogos tipo PrestaShop/OpenCart (el mismo
     motor que suelen usar sitios "hermanos" de distribución de scale models).
     Están marcados con "# AJUSTAR" — debes confirmarlos abriendo una página
     de producto real con F12 → pestaña "Elements" / "Inspeccionar".

  2. Un scraper con `requests` simple puede recibir error 403 (bloqueado).
     Por eso este archivo incluye DOS motores:
        - Motor A: requests + BeautifulSoup (rápido, pruébalo primero)
        - Motor B: Playwright (navegador real headless, esquiva bot-detection
          y es el único que te permite iniciar sesión si necesitas ver
          precios de distribuidor)

  3. Antes de scrapear en volumen, revisa https://www.modelcarswholesale.com/robots.txt
     y los Términos de Servicio del sitio. Como eres distribuidor con cuenta
     activa, lo más seguro (técnica y legalmente) es preguntarles si ofrecen
     un feed/export de catálogo (CSV, XML, API) — muchos mayoristas de este
     rubro sí lo tienen y te ahorra todo este trabajo.

--------------------------------------------------------------------------------
INSTALACIÓN (Windows, cmd o PowerShell):
--------------------------------------------------------------------------------
    python -m venv venv
    venv\\Scripts\\activate
    pip install requests beautifulsoup4 pandas openpyxl lxml
    pip install playwright
    playwright install chromium

--------------------------------------------------------------------------------
USO:
--------------------------------------------------------------------------------
    # Motor A (rápido, pruébalo primero)
    python scraper_modelcarswholesale.py --engine requests

    # Motor B (si el A da error 403 / bloqueo)
    python scraper_modelcarswholesale.py --engine playwright

    # Con login de distribuidor (variables de entorno, NUNCA hardcodeadas):
    #   set MCW_USER=tu_usuario
    #   set MCW_PASS=tu_clave
    python scraper_modelcarswholesale.py --engine playwright --login
================================================================================
"""

import os
import re
import sys
import time
import random
import logging
import argparse
from datetime import datetime
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ==============================================================================
# 1. CONFIGURACIÓN GENERAL
# ==============================================================================

BASE_URL = "https://www.modelcarswholesale.com"

# AJUSTAR: rutas reales de las secciones que quieres scrapear.
# Entra al menú del sitio, haz click en "Available" / "All Available" /
# "Special Offers" y copia la URL exacta que te da el navegador.

CATALOG_SECTIONS = {
    "available": "/search?keyword=&scale=all&type=available&country=all&page=1",
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "es-ES,es;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer": BASE_URL,
    "Connection": "keep-alive",
}

# Pausa aleatoria entre requests para no saturar el servidor ni parecer un bot
REQUEST_DELAY_RANGE = (1.8, 4.0)  # segundos

# Límite de seguridad para no entrar en un loop infinito si la paginación falla.
# Con catálogos de 1000+ productos, sube este número si ves que corta antes
# de llegar a la última página real (revisa cuántas páginas tiene cada sección).
MAX_PAGES_PER_SECTION = 300

OUTPUT_FILE = f"catalogo_modelcarswholesale_{datetime.now():%Y%m%d_%H%M}.xlsx"

# Cada cuántos productos se guarda un checkpoint (CSV) por si el proceso se
# corta a mitad de camino (caída de internet, bloqueo del sitio, etc.)
CHECKPOINT_CADA = 100
CHECKPOINT_FILE = "checkpoint_productos.csv"


def guardar_checkpoint(productos: list[dict]) -> None:
    """Guarda el progreso actual en CSV. Se sobreescribe cada vez (no acumula duplicados)."""
    if not productos:
        return
    pd.DataFrame(productos).to_csv(CHECKPOINT_FILE, index=False, encoding="utf-8-sig")
    logger.info(f"Checkpoint guardado: {len(productos)} productos en '{CHECKPOINT_FILE}'")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("mcw_scraper")


# ==============================================================================
# 2. (OPCIONAL) LOGIN PARA TARIFAS DE DISTRIBUIDOR — Motor A (requests.Session)
# ==============================================================================
def login_con_requests(session: requests.Session) -> bool:
    """
    Bloque preparado para iniciar sesión vía requests.Session, por si el sitio
    requiere login para ver precios/stock de mayorista.

    PASOS PARA CONFIGURARLO (hazlo una sola vez):
      1. Abre el sitio en Chrome, F12 → pestaña "Network", marca "Preserve log".
      2. Haz login manualmente con tu usuario y clave.
      3. En la lista de requests busca el POST hacia la URL de login
         (normalmente algo como /login o /connexion).
      4. Haz clic en esa request → pestaña "Payload" / "Form Data" y copia
         los nombres EXACTOS de los campos (ej. "email", "password", "_token").
      5. Reemplaza los valores de LOGIN_URL y las llaves del diccionario
         `payload` más abajo con esos nombres reales.

    Las credenciales se leen de variables de entorno (nunca las escribas
    directamente en el código ni las subas a un repositorio público).
    """
    LOGIN_URL = f"{BASE_URL}/login"  # <-- AJUSTAR a la URL real de login

    usuario = os.getenv("MCW_USER")
    clave = os.getenv("MCW_PASS")

    if not usuario or not clave:
        logger.info(
            "No se encontraron credenciales en variables de entorno "
            "(MCW_USER / MCW_PASS). Se continúa en modo público."
        )
        return False

    try:
        # Muchos sitios PrestaShop/OpenCart incluyen un token CSRF oculto en
        # el formulario de login. Si es el caso, hay que leerlo primero:
        pagina_login = session.get(LOGIN_URL, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(pagina_login.text, "lxml")

        token_input = soup.select_one('input[name="token"]')  # <-- AJUSTAR si aplica
        token = token_input["value"] if token_input else None

        payload = {
            "username": usuario,   # confirmado: input name="username" (no "email")
            "password": clave,
        }
        if token:
            payload["token"] = token  # <-- AJUSTAR nombre real del campo CSRF

        resp = session.post(LOGIN_URL, data=payload, headers=HEADERS, timeout=15)
        resp.raise_for_status()

        # Confirmado: cuando el login es exitoso aparece "Welcome" y "Log out"
        indicadores_login_ok = ["log out", "welcome"]
        texto_resp = resp.text.lower()

        if any(ind in texto_resp for ind in indicadores_login_ok):
            logger.info("Login exitoso (Motor A - requests).")
            return True

        logger.warning(
            "No se pudo confirmar el login con certeza. Revisa manualmente "
            "los selectores/indicadores en login_con_requests()."
        )
        return False

    except requests.RequestException as e:
        logger.error(f"Error durante el login (requests): {e}")
        return False


# ==============================================================================
# 3. MOTOR A — requests + BeautifulSoup
# ==============================================================================
def extraer_productos_de_pagina(html: str, url_pagina: str) -> list[dict]:
    """
    Recorre el HTML de una página de listado y extrae los datos de cada
    producto disponible. AJUSTAR los selectores según el HTML real del sitio.
    """
    productos = []
    soup = BeautifulSoup(html, "lxml")

    # AJUSTAR: selector del contenedor de cada tarjeta de producto.
    # Patrones típicos: "div.row product", "li.product", "div.product-miniature"
    tarjetas = soup.select("div.row.product")  # <-- AJUSTAR

    if not tarjetas:
        logger.warning(
            f"No se encontraron tarjetas de producto en {url_pagina}. "
            "Verifica el selector 'tarjetas' — puede que el sitio use "
            "JavaScript para renderizar el catálogo (usa --engine playwright)."
        )
        return productos

    for tarjeta in tarjetas:
        try:
            # --- Filtro de stock: saltar productos que NO están disponibles ---
            # Confirmado: div.availableText con texto "available" cuando hay stock.
            disponible_tag = tarjeta.select_one(".availableText")
            texto_disponible = disponible_tag.get_text(strip=True).lower() if disponible_tag else ""
            if "available" not in texto_disponible:
                continue

            # --- Nombre / descripción ---
            nombre_tag = tarjeta.select_one(".row-two.hidden-xs")
            nombre = nombre_tag.get_text(strip=True) if nombre_tag else ""

            # --- URL del producto y precio confirmado en .price ---
            enlace_tag = tarjeta.select_one(".thumb a")
            url_producto = urljoin(BASE_URL, enlace_tag["href"]) if enlace_tag and enlace_tag.has_attr("href") else None

            # --- Precio ---
            precio_tag = tarjeta.select_one(".price b")
            precio_texto = precio_tag.get_text(strip=True) if precio_tag else ""
            precio = limpiar_precio(precio_texto)

            # --- Imagen (confirmado: URL absoluta ya viene completa en src) ---
            img_tag = tarjeta.select_one(".thumb img")
            imagen_url = img_tag.get("src") if img_tag else None

            # --- Marca / Fabricante y Escala ---
            # Muchos catálogos de este rubro no separan marca/escala en campos
            # propios, sino que van dentro del nombre del producto
            # (ej. "SOLIDO 1:18 Nissan Skyline R34 Z-Tune").
            # Por eso se intentan extraer con expresiones regulares del nombre.
           
            caja_info = tarjeta.select_one(".scale")
            texto_info = caja_info.get_text(strip=True) if caja_info else ""
            
            # Extraemos los números para la escala (ej: "1:43")
            escala = extraer_escala(texto_info)
            
            # Para la marca, borramos los números de la escala y dejamos solo el texto (ej: "LOOKSMART")
            marca = re.sub(r"1[:/\-]\s?\d{2,3}", "", texto_info).strip() 
            
            # --- Trademark ---
            # Reemplaza ".trademark-info" con el nombre real que te dé el inspector
            caja_trademark = tarjeta.select_one(".trademark") 
            trademark = caja_trademark.get_text(strip=True) if caja_trademark else ""

            # --- Código ---
            caja_codigo = tarjeta.select_one('.row-three div[style*="width: 70%"] b')
            sku = caja_codigo.get_text(strip=True) if caja_codigo else ""

            productos.append(
                {
                    "SKU": sku,
                    "Marca": marca,
                    "Escala": escala,
                    "Nombre": nombre,
                    "Trademark": trademark,
                    "Precio": precio,
                    "Imagen_URL": imagen_url,
                    "URL_Producto": url_producto,
                }
            )

        except Exception as e:
            # Nunca dejamos que un solo producto mal formado tumbe todo el script
            logger.error(f"Error procesando una tarjeta de producto en {url_pagina}: {e}")
            continue

    return productos


def limpiar_precio(texto: str) -> str:
    """Extrae solo el valor numérico de un string de precio (ej. '€ 28.80' -> '28.80').
    El sitio usa punto como separador decimal (no coma), así que solo se
    quitan comas (posible separador de miles, ej. '1,234.56')."""
    if not texto:
        return ""
    match = re.search(r"[\d.,]+", texto)
    if not match:
        return ""
    numero = match.group(0).replace(",", "")
    try:
        return f"{float(numero):.2f}"
    except ValueError:
        return texto


def extraer_escala(nombre: str) -> str:
    """Busca patrones de escala (ej. 1:18, 1/43, 1:8, 1:2) dentro del texto."""
    # Cambiamos a (\d{1,3}) para que acepte desde 1 hasta 3 dígitos
    match = re.search(r"1[:/\-]\s?(\d{1,3})", nombre)
    return f"1:{match.group(1)}" if match else ""


def extraer_marca(nombre: str) -> str:
    """
    Intenta detectar la marca/fabricante como la primera palabra del nombre.
    AJUSTAR: si el sitio tiene un campo de marca separado (ej. un link a
    "/brand/solido"), es mucho más confiable extraerlo de ahí en vez de
    adivinar desde el texto.
    """
    if not nombre:
        return ""
    return nombre.split()[0]


def encontrar_url_pagina_siguiente(html: str, url_actual: str) -> str | None:
    """
    Busca el link de la página siguiente basándose en el símbolo de una sola flecha ('>').
    """
    soup = BeautifulSoup(html, "lxml")
    
    # Buscamos cualquier etiqueta 'a' cuyo texto contenga '>' pero NO '>>'
    siguiente = soup.find("a", string=lambda t: t and ">" in t and ">>" not in t)
    
    if siguiente and siguiente.has_attr("href"):
        return urljoin(url_actual, siguiente["href"])
        
    return None

def scrapear_con_requests(usar_login: bool = False) -> pd.DataFrame:
    """Motor A: recorre todas las secciones y páginas usando requests + BeautifulSoup."""
    session = requests.Session()
    session.headers.update(HEADERS)

    if usar_login:
        login_con_requests(session)

    todos_los_productos = []

    for nombre_seccion, ruta in CATALOG_SECTIONS.items():
        url = urljoin(BASE_URL, ruta)
        pagina_num = 1

        while url and pagina_num <= MAX_PAGES_PER_SECTION:
            logger.info(f"[{nombre_seccion}] Descargando página {pagina_num}: {url}")

            try:
                resp = session.get(url, timeout=20)
                resp.raise_for_status()
            except requests.RequestException as e:
                logger.error(f"Error al descargar {url}: {e}")
                break

            if resp.status_code == 403:
                logger.error(
                    "Bloqueo 403 detectado (bot-detection). "
                    "Vuelve a intentar con --engine playwright."
                )
                break

            productos_pagina = extraer_productos_de_pagina(resp.text, url)
            todos_los_productos.extend(productos_pagina)
            logger.info(f"  -> {len(productos_pagina)} productos extraídos en esta página.")
            logger.info(f"  -> Total acumulado: {len(todos_los_productos)} productos.")

            if len(todos_los_productos) % CHECKPOINT_CADA < len(productos_pagina):
                guardar_checkpoint(todos_los_productos)

            url_siguiente = encontrar_url_pagina_siguiente(resp.text, url)
            if not url_siguiente or url_siguiente == url:
                break
            url = url_siguiente
            pagina_num += 1

            # Pausa para respetar el servidor
            time.sleep(random.uniform(*REQUEST_DELAY_RANGE))

    return pd.DataFrame(todos_los_productos)


# ==============================================================================
# 4. MOTOR B — Playwright (recomendado si hay bot-detection, y necesario
#    si vas a iniciar sesión para ver precios de distribuidor)
# ==============================================================================
def scrapear_con_playwright(usar_login: bool = False, visible: bool = False, debug_port: int | None = None) -> pd.DataFrame:
    """
    Motor B: usa un navegador real para esquivar bot-detection y,
    opcionalmente, iniciar sesión antes de scrapear.
    Requiere: pip install playwright  &&  playwright install chromium

    Si debug_port se especifica, en vez de lanzar un navegador nuevo (que
    Cloudflare puede detectar como automatizado y dejarte en loop infinito
    de verificación), se conecta a un Chrome real que TÚ abriste a mano.
    Esto pasa el check de Cloudflare sin problema porque ese Chrome no tiene
    ninguna marca de automatización.
    """
    from playwright.sync_api import sync_playwright

    todos_los_productos = []

    
    with sync_playwright() as p:
        if debug_port:
            # Conectarse al Chrome "títere" que ya abriste a mano
            browser = p.chromium.connect_over_cdp(f"http://localhost:{debug_port}")
            context = browser.contexts[0]
            page = context.new_page()
        else:
            # Lanza su propio navegador. En modo visible (--visible) puedes
            # resolver tú mismo el check de seguridad de Cloudflare con un clic.
            browser = p.chromium.launch(
                headless=not visible,
                args=["--disable-blink-features=AutomationControlled"],
            )
            context = browser.new_context(
                user_agent=HEADERS["User-Agent"],
                locale="es-ES",
            )
            page = context.new_page()


        if usar_login:
            usuario = os.getenv("MCW_USER")
            clave = os.getenv("MCW_PASS")
            if usuario and clave:
                try:
                    page.goto(f"{BASE_URL}/login", timeout=30000)
                    if visible:
                        logger.info(
                            "Si aparece el check de Cloudflare, resuélvelo manualmente "
                            "en la ventana del navegador. Esperando hasta 90s..."
                        )
                        page.wait_for_selector('input[name="username"]', timeout=90000)
                    else:
                        page.wait_for_timeout(2000)
                    page.fill('input[name="username"]', usuario)
                    page.fill('input[name="password"]', clave)
                    page.click('button:has-text("Sign in")')
                    page.wait_for_load_state("domcontentloaded")
                    # Confirmar login real (no solo que la página cargó)
                    if "log out" in page.content().lower() or "welcome" in page.content().lower():
                        logger.info("Login realizado (Motor B - Playwright).")
                    else:
                        logger.warning("El login no se pudo confirmar (no aparece 'Log out'/'Welcome').")
                except Exception as e:
                    logger.error(f"No se pudo iniciar sesión con Playwright: {e}")
                    # DEBUG: guarda screenshot + HTML para ver qué le sirvió
                    # el sitio realmente al runner (posible bloqueo/Cloudflare)
                    try:
                        page.screenshot(path="debug_login.png", full_page=True)
                        with open("debug_login.html", "w", encoding="utf-8") as f:
                            f.write(page.content())
                        logger.info("Guardado debug_login.png / debug_login.html para diagnóstico.")
                    except Exception as e2:
                        logger.error(f"No se pudo guardar el debug: {e2}")
            else:
                logger.info(
                    "No se encontraron credenciales en variables de entorno "
                    "(MCW_USER / MCW_PASS). Se continúa en modo público."
                )

        for nombre_seccion, ruta in CATALOG_SECTIONS.items():
            url = urljoin(BASE_URL, ruta)
            pagina_num = 1

            while url and pagina_num <= MAX_PAGES_PER_SECTION:
                logger.info(f"[{nombre_seccion}] Cargando página {pagina_num}: {url}")

                try:
                    page.goto(url, timeout=10000)
                    page.wait_for_load_state("domcontentloaded")
                    page.wait_for_timeout(10000)
                except Exception as e:
                    logger.error(f"Error al cargar {url}: {e}")
                    break

                html = page.content()
                productos_pagina = extraer_productos_de_pagina(html, url)
                todos_los_productos.extend(productos_pagina)
                logger.info(f"  -> {len(productos_pagina)} productos extraídos en esta página.")

                url_siguiente = encontrar_url_pagina_siguiente(html, url)
                if not url_siguiente or url_siguiente == url:
                    break
                url = url_siguiente
                pagina_num += 1

                time.sleep(random.uniform(*REQUEST_DELAY_RANGE))

        browser.close()

    return pd.DataFrame(todos_los_productos)


# ==============================================================================
# 5. EXPORTACIÓN A EXCEL CON MINIATURAS (=IMAGE())
# ==============================================================================
def exportar_a_excel(df: pd.DataFrame, ruta_salida: str) -> None:
    """
    Genera el archivo .xlsx con una columna 'Foto' que usa la fórmula nativa
    =IMAGE("url") para mostrar la miniatura directamente en la celda.

    NOTA: la función =IMAGE() es nativa de Google Sheets y de Excel 365 /
    Excel para la Web (Microsoft la añadió como IMAGEN() / IMAGE() en 2023).
    Si abres el archivo en una versión de Excel de escritorio anterior a
    2023/365, la fórmula puede no evaluarse y solo verás el texto de la
    fórmula. En ese caso, sube el archivo a Google Sheets para verla.
    """
    if df.empty:
        logger.warning("El DataFrame está vacío — no se generará el Excel.")
        return

    df = df.copy()
    df["Foto"] = df["Imagen_URL"].apply(
        lambda u: f'=IMAGE("{u}")' if isinstance(u, str) and u else ""
    )

    columnas_orden = ["Foto", "Trademark", "Escala", "Nombre", "SKU", "codigo", "Precio"]
    columnas_orden = [c for c in columnas_orden if c in df.columns]
    df = df[columnas_orden]

    df.to_excel(ruta_salida, index=False, engine="openpyxl")

    # Ajustar ancho de columnas y alto de fila para que las miniaturas se vean bien
    wb = load_workbook(ruta_salida)
    ws = wb.active

    anchos = {"Foto": 14, "SKU": 16, "Marca": 16, "Escala": 10, "Nombre": 45, "Precio": 10}
    for idx, col in enumerate(columnas_orden, start=1):
        letra = get_column_letter(idx)
        ws.column_dimensions[letra].width = anchos.get(col, 20)

    for fila in range(2, ws.max_row + 1):
        ws.row_dimensions[fila].height = 60

    wb.save(ruta_salida)
    logger.info(f"Excel generado: {ruta_salida}  ({len(df)} productos)")


# ==============================================================================
# 6. MAIN
# ==============================================================================
def main():
    parser = argparse.ArgumentParser(description="Scraper de catálogo modelcarswholesale.com")
    parser.add_argument(
        "--engine",
        choices=["requests", "playwright"],
        default="playwright",
        help="Motor de scraping a usar. 'playwright' recomendado si hay bot-detection.",
    )
    parser.add_argument(
        "--login",
        action="store_true",
        help="Intenta iniciar sesión con las credenciales MCW_USER / MCW_PASS (variables de entorno).",
    )
    parser.add_argument(
        "--visible",
        action="store_true",
        help="Abre el navegador visible (no headless) para poder resolver checks de seguridad manualmente.",
    )
    parser.add_argument(
        "--debugport",
        type=int,
        default=None,
        help="Conecta a un Chrome real que abriste a mano con --remote-debugging-port=PUERTO (evita el loop de Cloudflare).",
    )
    parser.add_argument(
        "--maxpages",
        type=int,
        default=None,
        help="Límite de páginas a scrapear (para pruebas rápidas, ej. --maxpages 2). Sin esto, recorre el catálogo completo.",
    )
    args = parser.parse_args()

    logger.info(f"Iniciando scraping con motor: {args.engine}")

    if args.maxpages:
        global MAX_PAGES_PER_SECTION
        MAX_PAGES_PER_SECTION = args.maxpages
        logger.info(f"Modo prueba: limitado a {args.maxpages} página(s) por sección.")

    if args.engine == "requests":
        df = scrapear_con_requests(usar_login=args.login)
    else:
        df = scrapear_con_playwright(usar_login=args.login, visible=args.visible, debug_port=args.debugport)

    logger.info(f"Total de productos extraídos (todas las secciones): {len(df)}")

    exportar_a_excel(df, OUTPUT_FILE)

    # Nombre fijo (sin fecha) para que el workflow siempre sepa qué archivo
    # subir/commitear, y para que tu Google Sheet / Worker lo lean siempre
    # del mismo lugar.
    if not df.empty:
        df = asignar_categoria_envio(df, "excepciones_envio.csv")
        df = calcular_precios(df, "margenes_manuales.csv")

        # --- Archivo INTERNO (te quedas tú, nunca se sube a GitHub) ---
        columnas_internas = [
            "SKU", "Nombre", "Marca", "Trademark", "Escala", "Categoria_Envio",
            "FOB", "Precio_Envio", "Costo", "Margen", "Precio_Venta_Neto",
            "IGV", "Precio_Publico", "Ganancia",
        ]
        # Usar coma como separador decimal en los números (formato local)
        columnas_numericas_internas = [
            "FOB", "Precio_Envio", "Costo", "Margen",
            "Precio_Venta_Neto", "IGV", "Precio_Publico", "Ganancia",
        ]
        df_internos_out = df[columnas_internas].copy()
        for col in columnas_numericas_internas:
            if col in df_internos_out.columns:
                df_internos_out[col] = df_internos_out[col].map(
                    lambda v: f"{v:.2f}".replace(".", ",") if pd.notna(v) else v
                )
        df_internos_out.to_csv("precios_internos.csv", index=False, encoding="utf-8-sig")
        logger.info("CSV interno (costos/márgenes) guardado: precios_internos.csv")

        # --- Archivo PÚBLICO (este sí se sube, solo precio final) ---
        columnas_publicas = [
            "SKU", "Nombre", "Marca", "Trademark", "Escala",
            "Categoria_Envio", "Imagen_URL", "URL_Producto", "Precio_Publico",
        ]
        df_publico = df[columnas_publicas].rename(columns={"Precio_Publico": "Precio"})
        df_publico.to_csv("productos_actuales.csv", index=False, encoding="utf-8-sig")
        logger.info("CSV público (solo precio final) actualizado: productos_actuales.csv")

        subir_a_github()


def calcular_precios(df: pd.DataFrame, ruta_margenes: str) -> pd.DataFrame:
    """
    Calcula, para cada producto:
      Costo            = (FOB + Precio_Envio) * 1.04
      Margen           = 1.45 si Precio_Publico (probando ese margen) < 200, si no 1.25
      Precio_Venta_Neto= Costo * Margen
      Precio_Publico   = Precio_Venta_Neto * 1.18   (IGV 18%)
      Ganancia         = Precio_Venta_Neto - Costo

    Si existe margenes_manuales.csv (columnas SKU, Margen), ese margen
    fijo se usa en vez de la regla de los 200 — para casos especiales.
    Nunca se sobrescribe automáticamente: es tuyo, lo editas a mano.
    """
    df["FOB"] = pd.to_numeric(df["Precio"], errors="coerce").fillna(0)
    # Costo en soles: (FOB + envío) * 1.04 * 4 (tipo de cambio aprox. USD/EUR -> PEN)
    df["Costo"] = (df["FOB"] + df["Precio_Envio"]) * 1.04 * 4

    margenes_manuales = {}
    if os.path.exists(ruta_margenes):
        try:
            df_margenes = pd.read_csv(ruta_margenes, encoding="utf-8-sig")
            margenes_manuales = df_margenes.set_index("SKU")["Margen"].to_dict()
        except Exception as e:
            logger.warning(f"No se pudo leer margenes_manuales.csv: {e}")

    margenes, ventas_netas, igvs, publicos, ganancias = [], [], [], [], []
    for _, fila in df.iterrows():
        sku = fila.get("SKU")
        costo = fila["Costo"]

        if sku in margenes_manuales:
            margen = margenes_manuales[sku]
        else:
            # Probar primero el margen alto (1.45); si el precio público
            # resultante se pasa de 200, usar el margen bajo (1.25).
            precio_prueba = costo * 1.45 * 1.18
            margen = 1.45 if precio_prueba < 200 else 1.25

        venta_neta = costo * margen
        igv = venta_neta * 0.18
        publico = venta_neta + igv
        ganancia = venta_neta - costo

        margenes.append(margen)
        ventas_netas.append(round(venta_neta, 2))
        igvs.append(round(igv, 2))
        publicos.append(round(publico, 2))
        ganancias.append(round(ganancia, 2))

    df["Margen"] = margenes
    df["Precio_Venta_Neto"] = ventas_netas
    df["IGV"] = igvs
    df["Precio_Publico"] = publicos
    df["Ganancia"] = ganancias

    # --- Precio final manual (gana siempre sobre el cálculo automático) ---
    # Archivo aparte que TÚ mantienes: precios_finales_manuales.csv
    #   SKU,Precio_Publico
    #   18-16023R,45.90
    # El scraper nunca lo sobrescribe. Si un SKU está ahí, ese precio se usa
    # tal cual (recalculando venta neta/IGV/ganancia hacia atrás para que el
    # desglose interno siga siendo consistente).
    ruta_precios_finales = "precios_finales_manuales.csv"
    if os.path.exists(ruta_precios_finales):
        try:
            df_precios_finales = pd.read_csv(ruta_precios_finales, encoding="utf-8-sig")
            overrides = df_precios_finales.set_index("SKU")["Precio_Publico"].to_dict()
            aplicados = 0
            for i, fila in df.iterrows():
                sku = fila.get("SKU")
                if sku in overrides:
                    publico_manual = float(overrides[sku])
                    venta_neta_manual = publico_manual / 1.18
                    df.at[i, "Precio_Publico"] = round(publico_manual, 2)
                    df.at[i, "Precio_Venta_Neto"] = round(venta_neta_manual, 2)
                    df.at[i, "IGV"] = round(publico_manual - venta_neta_manual, 2)
                    df.at[i, "Ganancia"] = round(venta_neta_manual - fila["Costo"], 2)
                    aplicados += 1
            logger.info(f"Precios finales manuales aplicados: {aplicados} producto(s).")
        except Exception as e:
            logger.warning(f"No se pudo leer precios_finales_manuales.csv: {e}")

    return df


# Palabras clave (en inglés, porque el catálogo del proveedor está en
# inglés) para detectar vehículos grandes y aviones (envío distinto al
# de un auto estándar). AJUSTAR/agregar las que falten.
PALABRAS_CLAVE_CAMION = [
    "truck", "bus", "coach", "tractor", "trailer", "semi",
    "van", "fire engine", "firetruck", "pickup", "crane",
    "bulldozer", "excavator", "cement mixer", "tanker",
]

PALABRAS_CLAVE_AVION = [
    "aircraft", "airplane", "plane", "jet", "helicopter",
    "airliner", "fighter", "bomber", "biplane", "glider",
]

PALABRAS_CLAVE_MINIATURA = [
    "helmet", "figure", "figurine", "doll", "statue", "bust",
    "keychain", "keyring", "magnet", "pin badge", "mini", "miniature",
]

PALABRAS_CLAVE_MOTO = [
    "motorcycle", "motorbike", "scooter", "moped", "quad bike", "atv",
]

# --- Tablas de envío por escala, según categoría ---
# Cada tupla es (denominador_min, denominador_max, precio). Un denominador
# más GRANDE significa un modelo más CHICO (ej. 1/43 es más chico que 1/18).
# None en denominador_max significa "sin límite superior" (denom >= min);
# None en denominador_min significa "sin límite inferior" (denom <= max).
TABLA_ENVIO_COCHES = [
    (43, None, 2),   # 1/43 y más pequeños
    (24, 42, 3),     # entre 1/42 y 1/24
    (19, 23, 5),     # entre 1/23 y 1/19
    (18, 18, 7),     # 1/18
    (12, 17, 10),    # entre 1/17 y 1/12
    (None, 10, 14),  # 1/10 y más grandes
]

TABLA_ENVIO_CAMIONES = [
    (43, None, 5),
    (24, 42, 10),
    (19, 23, 14),
    (18, 18, 7),
    (12, 17, 20),
    (None, 10, 30),
]

TABLA_ENVIO_AVIONES = [
    (43, None, 110),
    (24, 42, 14),
    (19, 23, 20),
    (18, 18, 7),
    (12, 17, 30),
    (None, 10, 50),
]

TABLA_ENVIO_MOTOS = [
    (18, 18, 2),
    (12, 12, 3),
    (6, 6, 10),
]

TABLA_ENVIO_MINIATURA = [
    (5, 5, 2),
    (4, 4, 3),
    (2, 2, 5),
]

TABLAS_ENVIO_POR_CATEGORIA = {
    "estandar": TABLA_ENVIO_COCHES,
    "camion": TABLA_ENVIO_CAMIONES,
    "avion": TABLA_ENVIO_AVIONES,
    "moto": TABLA_ENVIO_MOTOS,
    "miniatura": TABLA_ENVIO_MINIATURA,
}


def extraer_denominador_escala(escala: str):
    """De un texto de escala tipo '1/24' extrae el denominador (24) como int.
    Devuelve None si no se pudo interpretar."""
    if not escala:
        return None
    match = re.search(r"1\s*/\s*(\d+)", str(escala))
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def precio_envio_por_escala(categoria: str, escala: str) -> float:
    """
    Busca el precio de envío según la tabla de la categoría y el
    denominador de la escala. Si el denominador cae exactamente dentro de
    un rango de la tabla, usa ese precio. Si cae en un hueco entre dos
    rangos (denominador no cubierto), promedia el precio del rango
    inmediatamente más chico (denom mayor) y el inmediatamente más
    grande (denom menor). Si no hay escala reconocible, usa el precio
    "estándar" por defecto de esa categoría (rango 1/18) como respaldo.
    """
    tabla = TABLAS_ENVIO_POR_CATEGORIA.get(categoria, TABLA_ENVIO_COCHES)
    denom = extraer_denominador_escala(escala)

    if denom is None:
        # Sin escala reconocible: usar un valor intermedio de la tabla
        precios_tabla = [p for (_, _, p) in tabla]
        return sum(precios_tabla) / len(precios_tabla)

    # 1) Coincidencia exacta dentro de algún rango
    for lo, hi, precio in tabla:
        dentro_min = (lo is None) or (denom >= lo)
        dentro_max = (hi is None) or (denom <= hi)
        if dentro_min and dentro_max:
            return precio

    # 2) No hay coincidencia exacta.
    #    Para "moto" y "miniatura": promediar entre el rango más cercano
    #    por arriba y por abajo (como se pidió).
    #    Para el resto (estandar, camion, avion): NO promediar, usar el
    #    precio del rango cuyo límite esté más cerca del denominador buscado.
    vecino_mayor_denom = None   # rango con denom > el buscado (modelo más chico), el más cercano
    vecino_menor_denom = None   # rango con denom < el buscado (modelo más grande), el más cercano

    for lo, hi, precio in tabla:
        # Extremo "chico" (denom alto) de este rango
        extremo_alto = hi if hi is not None else float("inf")
        extremo_bajo = lo if lo is not None else 0

        if extremo_bajo > denom:
            if vecino_mayor_denom is None or extremo_bajo < vecino_mayor_denom[0]:
                vecino_mayor_denom = (extremo_bajo, precio)
        if extremo_alto < denom:
            if vecino_menor_denom is None or extremo_alto > vecino_menor_denom[0]:
                vecino_menor_denom = (extremo_alto, precio)

    if categoria in ("moto", "miniatura"):
        if vecino_mayor_denom and vecino_menor_denom:
            return (vecino_mayor_denom[1] + vecino_menor_denom[1]) / 2
        elif vecino_mayor_denom:
            return vecino_mayor_denom[1]
        elif vecino_menor_denom:
            return vecino_menor_denom[1]
        else:
            return tabla[0][2]
    else:
        # Sin promedio: elegir el límite más cercano al denominador buscado
        if vecino_mayor_denom and vecino_menor_denom:
            dist_mayor = abs(vecino_mayor_denom[0] - denom)
            dist_menor = abs(vecino_menor_denom[0] - denom)
            return vecino_mayor_denom[1] if dist_mayor <= dist_menor else vecino_menor_denom[1]
        elif vecino_mayor_denom:
            return vecino_mayor_denom[1]
        elif vecino_menor_denom:
            return vecino_menor_denom[1]
        else:
            return tabla[0][2]


def clasificar_categoria_envio(nombre: str) -> str:
    """
    Detecta automáticamente la categoría de envío según palabras clave
    (en inglés) en el nombre/descripción del producto:
      - "miniatura": cascos, figuras, muñequitos, llaveros, etc.
      - "avion": aviones, helicópteros, etc.
      - "camion": camiones, buses, tractores, etc.
      - "moto": motos, scooters, cuatrimotos, etc.
      - "estandar": autos normales (todo lo demás).
    """
    texto = (nombre or "").lower()
    if any(palabra in texto for palabra in PALABRAS_CLAVE_MINIATURA):
        return "miniatura"
    if any(palabra in texto for palabra in PALABRAS_CLAVE_AVION):
        return "avion"
    if any(palabra in texto for palabra in PALABRAS_CLAVE_CAMION):
        return "camion"
    if any(palabra in texto for palabra in PALABRAS_CLAVE_MOTO):
        return "moto"
    return "estandar"


def asignar_categoria_envio(df: pd.DataFrame, ruta_excepciones: str) -> pd.DataFrame:
    """
    1. Clasifica todos los productos automáticamente por palabras clave.
    2. Si existe excepciones_envio.csv (con columnas SKU, Categoria_Envio),
       esas asignaciones manuales ganan siempre sobre la detección
       automática — para los casos raros que el detector no identifica bien.

    excepciones_envio.csv es un archivo que TÚ mantienes a mano (el
    scraper nunca lo sobrescribe), con formato:
        SKU,Categoria_Envio
        18-16023R,camion
    """
    df["Categoria_Envio"] = df["Nombre"].apply(clasificar_categoria_envio)
    df["Precio_Envio"] = df.apply(
        lambda fila: precio_envio_por_escala(fila["Categoria_Envio"], fila.get("Escala")),
        axis=1,
    )

    if os.path.exists(ruta_excepciones):
        try:
            df_excepciones = pd.read_csv(ruta_excepciones, encoding="utf-8-sig")
            excepciones = df_excepciones.set_index("SKU")["Categoria_Envio"].to_dict()
            aplicadas = 0
            for i, fila in df.iterrows():
                sku = fila.get("SKU")
                if sku in excepciones:
                    df.at[i, "Categoria_Envio"] = excepciones[sku]
                    df.at[i, "Precio_Envio"] = precio_envio_por_escala(
                        excepciones[sku], fila.get("Escala")
                    )
                    aplicadas += 1
            logger.info(f"Excepciones de envío aplicadas: {aplicadas} producto(s).")
        except Exception as e:
            logger.warning(f"No se pudo leer excepciones_envio.csv: {e}")
    else:
        logger.info(
            "No existe excepciones_envio.csv (opcional) — usando solo "
            "detección automática por palabras clave."
        )

    return df


def respetar_precios_manuales(df_nuevo: pd.DataFrame, ruta_csv_anterior: str) -> pd.DataFrame:
    """
    Si ya existe un productos_actuales.csv de una corrida anterior, conserva
    el precio que estaba ahí para cada SKU (por si lo editaste a mano en
    Excel), en vez de sobrescribirlo con el precio recién scrapeado.
    Los productos nuevos sí usan el precio del proveedor (no hay nada previo
    que conservar).
    """
    if not os.path.exists(ruta_csv_anterior):
        return df_nuevo

    try:
        df_anterior = pd.read_csv(ruta_csv_anterior, encoding="utf-8-sig")
    except Exception as e:
        logger.warning(f"No se pudo leer el CSV anterior para conservar precios: {e}")
        return df_nuevo

    if "SKU" not in df_anterior.columns or "Precio" not in df_anterior.columns:
        return df_nuevo

    precios_anteriores = df_anterior.set_index("SKU")["Precio"].to_dict()

    conservados = 0
    for i, fila in df_nuevo.iterrows():
        sku = fila.get("SKU")
        if sku in precios_anteriores:
            df_nuevo.at[i, "Precio"] = precios_anteriores[sku]
            conservados += 1

    logger.info(f"Precios manuales conservados: {conservados} producto(s).")
    return df_nuevo


def subir_a_github():
    """
    Sube productos_actuales.csv al repo de GitHub automáticamente.
    Requiere que esta carpeta sea un repo git ya clonado y con credenciales
    configuradas (ver instrucciones de configuración inicial más abajo).
    """
    import subprocess

    try:
        subprocess.run(["git", "add", "productos_actuales.csv"], check=True)
        resultado = subprocess.run(
            ["git", "commit", "-m", "Actualizacion de catalogo"],
            capture_output=True, text=True,
        )
        if resultado.returncode != 0:
            logger.info("Nada nuevo que commitear (el catálogo no cambió).")
            return
        subprocess.run(["git", "push"], check=True)
        logger.info("CSV subido a GitHub correctamente.")
    except subprocess.CalledProcessError as e:
        logger.error(f"No se pudo subir a GitHub automáticamente: {e}")


if __name__ == "__main__":
    main()
